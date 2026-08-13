#!/usr/bin/env python3
"""
price_anchor.py — 报价单 UUID 锚点：生成隐藏列 + 价格回写（行序无关）

为什么需要它：
  报价单发给客户后，客户常会插行、删行、调顺序、加分组标题。
  若按"第 N 行价格写回第 N 条数据"回写，行序一变价格就全错位。
  本脚本在报价单里埋一列隐藏的 __item_uuid__，回写时靠 UUID 精确定位每一行，
  用户怎么编辑 Excel 都不影响（设计来源：trade-pipeline 的真实生产 bug）。

两个命令：
  stamp  给报价单写入/刷新隐藏 UUID 列（按品名匹配 order.json 的 item_uuid）
  update 从填好单价的报价单回写单价到 order.json（按 UUID 定位，重算 amount）

用法：
  python price_anchor.py stamp  <报价单.xlsx> <order.json>
  python price_anchor.py update <报价单.xlsx> <order.json>

依赖：openpyxl、标准库。
"""
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少 openpyxl，请先: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

UUID_HEADER = "__item_uuid__"

# 品名列 / 单价列 的列头关键词（小写匹配）
DESC_KEYWORDS = ["品名", "description", "item", "product", "наименование", "описание", "名称", "货品"]
PRICE_KEYWORDS = ["单价", "unit price", "цена", "price"]


def _norm(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").lower().strip())


def _find_header_row(ws, max_scan=30):
    """找表头行：含任意品名列关键词的行。"""
    for r in range(1, min(ws.max_row + 1, max_scan)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "")
            if any(k in v.lower() for k in DESC_KEYWORDS):
                return r
    return None


def _find_col(ws, header_row, keywords, exclude=()):
    """在表头行找含关键词的列（排除 exclude 列号）。"""
    for c in range(1, ws.max_column + 1):
        if c in exclude:
            continue
        v = str(ws.cell(header_row, c).value or "").lower()
        if any(k in v for k in keywords):
            return c
    return None


def _find_uuid_col(ws, max_scan=30):
    for r in range(1, min(ws.max_row + 1, max_scan)):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip() == UUID_HEADER:
                return c
    return None


def _find_price_col(ws, header_row, uuid_col):
    """找单价列，优先"单价"（排除总价/金额列），找不到回退含 price 的列。"""
    # 精确"单价/unit price/цена"优先
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").lower()
        if ("单价" in v) or ("unit price" in v) or ("цена" in v):
            if "总" not in v and "total" not in v and "amount" not in v and "сумма" not in v:
                return c
    # 回退：任意含 price 的列（排除 total/amount）
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").lower()
        if "price" in v or "цена" in v:
            if "total" not in v and "amount" not in v and "сумма" not in v:
                return c
    return None


# ── stamp：写隐藏 UUID 列 ─────────────────────────────────────────


def stamp(quote_path: str, order_path: str) -> dict:
    order = json.loads(Path(order_path).read_text(encoding="utf-8"))
    items = order.get("items", [])

    wb = load_workbook(quote_path)
    ws = wb.active

    header_row = _find_header_row(ws)
    if header_row is None:
        raise SystemExit("✗ 未找到报价单表头行（无品名列），请确认报价单结构。")

    desc_col = _find_col(ws, header_row, DESC_KEYWORDS)
    if desc_col is None:
        raise SystemExit("✗ 未找到品名列。")

    # 已有 UUID 列则沿用，否则新建在最右侧
    uuid_col = _find_uuid_col(ws)
    if uuid_col is None:
        uuid_col = ws.max_column + 1
        ws.cell(header_row, uuid_col).value = UUID_HEADER

    # 隐藏 UUID 列
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(uuid_col)].hidden = True

    # 品名 → item 匹配（规范化后相等；重复品名按出现顺序分配）
    desc_to_items = {}
    for it in items:
        for d in (it.get("description_en", ""), it.get("description_cn", ""), it.get("description_ru", "")):
            if d:
                desc_to_items.setdefault(_norm(d), []).append(it)

    used = {}
    matched = 0
    total_data_rows = 0
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(r, desc_col).value
        if desc is None or not str(desc).strip():
            continue  # 跳过空行/分组标题行
        total_data_rows += 1
        key = _norm(desc)
        cands = desc_to_items.get(key, [])
        if not cands:
            continue
        it = cands[min(used.get(key, 0), len(cands) - 1)]
        used[key] = used.get(key, 0) + 1
        ws.cell(r, uuid_col).value = it["item_uuid"]
        matched += 1

    wb.save(quote_path)
    return {"header_row": header_row, "uuid_col": uuid_col, "desc_col": desc_col,
            "matched": matched, "total_data_rows": total_data_rows, "total_items": len(items)}


# ── update：回写单价 ──────────────────────────────────────────────


def update(quote_path: str, order_path: str) -> dict:
    order = json.loads(Path(order_path).read_text(encoding="utf-8"))
    items = order.get("items", [])
    by_uuid = {it["item_uuid"]: it for it in items if it.get("item_uuid")}

    wb = load_workbook(quote_path, data_only=True)
    ws = wb.active

    uuid_col = _find_uuid_col(ws)
    if uuid_col is None:
        raise SystemExit(f"✗ 报价单中未找到隐藏列 '{UUID_HEADER}'。请先执行 stamp。")

    header_row = _find_header_row(ws)
    price_col = _find_price_col(ws, header_row, uuid_col)
    if price_col is None:
        raise SystemExit("✗ 未找到单价列（列头需含「单价/Unit Price/цена」）。")

    updated = 0
    errors = []
    warnings = []
    for r in range(header_row + 1, ws.max_row + 1):
        uid = ws.cell(r, uuid_col).value
        if uid is None or str(uid).strip() in ("", UUID_HEADER):
            continue
        uid = str(uid).strip()
        if uid not in by_uuid:
            warnings.append(f"行 {r} 的 uuid '{uid}' 不在 order.json 中（可能已删除）")
            continue
        price = ws.cell(r, price_col).value
        if price is None or str(price).strip() == "":
            warnings.append(f"item {uid} (行 {r}) 单价为空，跳过")
            continue
        try:
            pv = float(price)
        except (ValueError, TypeError):
            errors.append(f"item {uid} (行 {r}) 单价格式非法: {price!r}")
            continue
        if pv < 0:
            errors.append(f"item {uid} (行 {r}) 单价为负: {pv}")
            continue
        by_uuid[uid]["unit_price"] = pv
        q = by_uuid[uid].get("quantity")
        if q is not None:
            by_uuid[uid]["amount"] = round(q * pv, 2)
        updated += 1

    if errors:
        print("✗ 存在错误，不落盘：")
        for e in errors:
            print(f"  {e}")
        return {"updated": 0, "errors": errors, "warnings": warnings}

    Path(order_path).write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"updated": updated, "errors": errors, "warnings": warnings}


# ── CLI ───────────────────────────────────────────────────────────


def _main(argv: list[str]) -> int:
    if len(argv) < 4:
        print(__doc__)
        return 2
    cmd, quote_path, order_path = argv[1], argv[2], argv[3]

    if cmd == "stamp":
        r = stamp(quote_path, order_path)
        print(f"✅ 已写隐藏 UUID 列（列 {r['uuid_col']}，表头行 {r['header_row']}）")
        print(f"   品名匹配 {r['matched']}/{r['total_data_rows']} 行，order.json 共 {r['total_items']} 条")
        if r["matched"] < r["total_items"]:
            print("   ⚠ 有 item 未匹配到报价单行，请核对品名是否一致")
        return 0

    if cmd == "update":
        r = update(quote_path, order_path)
        print(f"✅ 已回写 {r['updated']} 条单价到 {order_path}（amount 已重算）")
        for w in r["warnings"]:
            print(f"   ⚠ {w}")
        return 0 if not r["errors"] else 1

    print(__doc__)
    return 2


if __name__ == "__main__":
    raise SystemExit(_main(sys.argv))
