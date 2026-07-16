"""
Trade AI Assistant — 管理表格模板生成器。

生成 10 张结构化 Excel 管理表格，覆盖外贸全流程：
选品策略 → 客户画像 → 市场分析 → 渠道策略 → 内容日历
→ 跟进时间表 → 管线看板 → KPI 追踪 → 竞品分析 → 文化适配
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 颜色常量 ─────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SUBHEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
_BODY_FONT = Font(name="微软雅黑", size=9)
_NOTE_FONT = Font(name="微软雅黑", size=8, color="808080")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_row(ws, row: int, max_col: int, *, alt: bool = False):
    fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") if alt else PatternFill()
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        cell.alignment = _WRAP
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = fill


def _add_title(ws, title: str, row: int, cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = f"【{title}】"
    cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1 — 选品与产品策略
# ═══════════════════════════════════════════════════════════════════════════

def _build_selection(ws: Workbook):
    ws_ = ws.active if ws.active else ws.create_sheet()
    ws_.title = "1-选品策略"
    ws_.sheet_properties.tabColor = "1F4E79"

    headers = [
        "编号", "产品名称/SKU", "目标市场/区域", "选品标准(必备)",
        "选品标准(加分)", "核心卖点", "差异化定位描述",
        "市场机会评估", "认证要求", "建议优先级",
    ]
    _add_title(ws_, "选品与产品策略规划表", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    # 示例行
    samples = [
        [1, "例：EC 智能电机", "欧洲 / 德国", "CE 认证, UL 认证",
         "能效等级 IE4+", "比同类节能 20%", "专为数据中心设计的 EC 智能电机",
         "欧盟绿色新政推动高能效电机需求增长", "CE, RoHS, REACH", "P0"],
    ]
    for ri, row in enumerate(samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    # 空模板行
    for ri in range(5, 25):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    ws_.column_dimensions["A"].width = 6
    for c in range(2, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 22


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2 — 客户画像分析表
# ═══════════════════════════════════════════════════════════════════════════

def _build_persona(ws: Workbook):
    ws_ = ws.create_sheet("2-客户画像")
    ws_.sheet_properties.tabColor = "2E75B6"

    headers = [
        "客户类型", "决策角色", "岗位职责/KPI", "核心痛点(3-5)",
        "关注的 Feature", "Advantage(比较优势)", "Benefit(按角色翻译)",
        "沟通 Hook", "建议渠道", "常见反对意见与应对",
    ]
    _add_title(ws_, "客户画像与角色分层分析表", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    roles = [
        ["例：品牌商/分销商", "采购经理", "控制采购成本、确保供应链稳定",
         "价格波动大、交期不可控、供应商替换成本高",
         "有竞争力的定价、稳定交期", "同行平均交期 25天，我们 15天",
         "降低库存成本 30%、减少缺货风险",
         "「同品类客户 X 与您合作后库存周转率提升了 35%」",
         "Email / LinkedIn",
         "「你们价格比现有供应商高」→ 对比 TCO(总拥有成本)"],
        ["", "技术/工程师", "产品性能达标准、认证齐全",
         "质量问题、认证不匹配、技术参数不符",
         "完整的认证体系、详细规格书", "CE+UL 双认证、提供第三方检测报告",
         "降低验证风险、加速产品导入",
         "「我们为贵行业的标杆企业 Y 提供了定制方案」", "Email / 视频会议", ""],
        ["", "高层决策者", "ROI、市场份额、竞争优势",
         "看不清供应商价值、替换风险大",
         "成功案例、ROI 数据", "同行业客户平均 ROI 提升 25%",
         "战略价值：加速新品上市、提升竞争力",
         "「一个季度看到 ROI — 已有 3 家同行选择了我们」",
         "LinkedIn InMail / 高层会面", ""],
    ]
    for ri, row in enumerate(roles, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(7, 25):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    ws_.column_dimensions["A"].width = 14
    for c in range(2, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 24


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 3 — 市场分析作战地图
# ═══════════════════════════════════════════════════════════════════════════

def _build_market(ws: Workbook):
    ws_ = ws.create_sheet("3-市场分析")
    ws_.sheet_properties.tabColor = "548235"

    headers = [
        "目标国家", "产品名称", "HS 编码", "强制认证",
        "推荐认证", "进口关税(%)", "价值词(10个)",
        "高意图搜索词", "长尾问题词", "邮件主题行Hook",
        "LinkedIn 开场白", "市场验证行动",
    ]
    _add_title(ws_, "市场分析作战地图", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    samples = [
        ["德国", "EC 智能电机", "8501.40", "CE, RoHS",
         "TUV、VDE", "2.5%",
         "reliable, cost-effective, turnkey, innovative, certified",
         "EC motor Germany, energy efficient motor EU",
         "how to choose EC motor supplier, energy saving motor EU standards",
         "「Reducing your motor energy cost by 20%」",
         "「Saw your interest in energy-efficient solutions at X展会」",
         "LinkedIn 触达当地经销商、测试 Google Ads"],
    ]
    for ri, row in enumerate(samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(5, 15):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 22


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 4 — 渠道策略矩阵
# ═══════════════════════════════════════════════════════════════════════════

def _build_channel(ws: Workbook):
    ws_ = ws.create_sheet("4-渠道策略")
    ws_.sheet_properties.tabColor = "BF8F00"

    headers = [
        "客户类型", "主要渠道", "备用渠道", "选择理由",
        "内容方向", "联系方式模板",
    ]
    _add_title(ws_, "渠道策略与触达矩阵", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    samples = [
        ["分销商/进口商", "LinkedIn + Email", "行业展会",
         "分销商活跃在 LinkedIn,B2B 采购习惯用 Email",
         "产品优势+渠道政策+成功案例",
         "「We are looking for distributors in [Country] for [Product]」"],
        ["品牌商/OEM", "LinkedIn InMail", "Email + 电话跟进",
         "决策者在 LinkedIn 活跃度高",
         "定制能力+品质认证+研发实力",
         "「We support OEM/ODM for global brands in [Industry]」"],
    ]
    for ri, row in enumerate(samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(6, 20):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 28


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 5 — AIDA 内容日历
# ═══════════════════════════════════════════════════════════════════════════

def _build_content(ws: Workbook):
    ws_ = ws.create_sheet("5-内容日历")
    ws_.sheet_properties.tabColor = "C00000"

    headers = [
        "日期/周", "AIDA 阶段", "渠道", "内容主题",
        "内容格式", "Hook 类型", "CTA", "目标受众",
    ]
    _add_title(ws_, "AIDA 内容日历（建议每周 ≥3 条）", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    samples = [
        ["第 1 周", "Awareness", "LinkedIn", "行业趋势分析：XX 市场 2026 展望",
         "图文帖", "数据引用", "评论互动", "目标行业采购经理"],
        ["第 1 周", "Awareness", "Facebook", "工厂产线实拍视频",
         "短视频(30s)", "视觉冲击", "私信咨询", "潜在客户"],
        ["第 2 周", "Interest", "EDM", "客户案例：如何帮 X 降低 20% 成本",
         "案例邮件", "客户故事", "预约视频会议", "已互动客户"],
    ]
    for ri, row in enumerate(samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(7, 55):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 22


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 6 — 30 天跟进时间表
# ═══════════════════════════════════════════════════════════════════════════

def _build_followup(ws: Workbook):
    ws_ = ws.create_sheet("6-跟进时间表")
    ws_.sheet_properties.tabColor = "7030A0"

    headers = [
        "Day", "客户旅程阶段", "动作类型", "具体行动",
        "模板/话术方向", "推进目标", "无回应下一步",
    ]
    _add_title(ws_, "30 天跟进时间表", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    samples = [
        ["Day 1", "Awareness", "开发信", "首次触达,个性化开发信",
         "行业洞察+产品价值", "获得回复", "—"],
        ["Day 3", "Awareness", "跟进邮件", "无回复跟进",
         "补充信息+增加 Hook 强度", "获得回复", "换渠道 LinkedIn"],
        ["Day 7", "Interest", "LinkedIn", "LinkedIn 互动+InMail",
         "分享行业文章并提问", "建立对话", "电话跟进"],
        ["Day 14", "Consideration", "视频会议", "产品演示/方案介绍",
         "针对客户痛点展示方案", "推进到报价阶段", "发资料包"],
        ["Day 21", "Decision", "报价跟进", "报价+条款确认",
         "报价单+差异化总结", "收到确认", "发送补充案例"],
        ["Day 30", "Decision", "推进", "价格谈判",
         "合作价值+ROI 重申", "成交", "降级至培育序列"],
    ]
    for ri, row in enumerate(samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(10, 35):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    ws_.column_dimensions["A"].width = 8
    for c in range(2, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 26


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 7 — 管线健康度看板
# ═══════════════════════════════════════════════════════════════════════════

def _build_pipeline(ws: Workbook):
    ws_ = ws.create_sheet("7-管线看板")
    ws_.sheet_properties.tabColor = "00B050"

    headers = [
        "客户名称", "国家", "客户类型", "当前阶段",
        "状态", "最近动作日期", "最近动作", "本周建议动作",
        "风险评估", "优先级",
    ]
    _add_title(ws_, "客户管线健康度看板", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    for ri in range(4, 30):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    ws_.column_dimensions["A"].width = 16
    for c in range(2, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 20


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 8 — KPI 追踪表
# ═══════════════════════════════════════════════════════════════════════════

def _build_kpi(ws: Workbook):
    ws_ = ws.create_sheet("8-KPI追踪")
    ws_.sheet_properties.tabColor = "FF0000"

    headers = [
        "KPI 指标", "定义", "目标值", "本周实际", "本月实际",
        "趋势(↑↓→)", "复盘频率", "改善措施",
    ]
    _add_title(ws_, "KPI 追踪与复盘表", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    kpi_samples = [
        ["有效接触率", "回复/开发信发出数", "25-40%", "", "", "", "每周", ""],
        ["样品转化率", "要样品/报价客户中成交", "15-25%", "", "", "", "每周", ""],
        ["视频会议预约率", "会议数/有效接触数", "10-20%", "", "", "", "每周", ""],
        ["报价接受率", "接受报价/发出报价", "20-35%", "", "", "", "双周", ""],
        ["成交率", "成交数/总线索数", "5-10%", "", "", "", "每月", ""],
    ]
    for ri, row in enumerate(kpi_samples, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(9, 20):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 20


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 9 — 竞品差异分析
# ═══════════════════════════════════════════════════════════════════════════

def _build_competitor(ws: Workbook):
    ws_ = ws.create_sheet("9-竞品分析")
    ws_.sheet_properties.tabColor = "808080"

    headers = [
        "对比维度", "我方", "竞品 A", "竞品 B",
        "我方优势", "差距/待改进",
    ]
    _add_title(ws_, "竞品差异分析表", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    dims = [
        ["价格定位", "", "", "", "", ""],
        ["交期", "", "", "", "", ""],
        ["认证资质", "", "", "", "", ""],
        ["最小起订量", "", "", "", "", ""],
        ["定制能力", "", "", "", "", ""],
        ["售后服务", "", "", "", "", ""],
        ["品牌知名度", "", "", "", "", ""],
        ["目标市场", "", "", "", "", ""],
    ]
    for ri, row in enumerate(dims, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 24


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 10 — 文化适配
# ═══════════════════════════════════════════════════════════════════════════

def _build_culture(ws: Workbook):
    ws_ = ws.create_sheet("10-文化适配")
    ws_.sheet_properties.tabColor = "ED7D31"

    headers = [
        "目标国/地区", "商业文化特点", "沟通偏好",
        "谈判风格", "付款习惯", "禁忌/注意事项",
        "送礼/社交礼仪", "最佳触达时段",
    ]
    _add_title(ws_, "跨文化商业适配指南", 1, len(headers))
    _add_title(ws_, f"生成日期：{date.today().isoformat()}", 2, len(headers))

    for i, h in enumerate(headers, 1):
        ws_.cell(row=3, column=i, value=h)
    _style_header(ws_, 3, len(headers))

    cultures = [
        ["德国", "严谨、重计划、对品质要求极高",
         "正式 Email 为首选，避免 WhatsApp 初次联系",
         "直截了当、以数据和事实说话",
         "TT 为主，大额 L/C",
         "避免过于热情的营销语言",
         "商务礼品不宜贵重，精致有品质感即可",
         "工作时间 9:00-17:00 CET"],
        ["中东(海湾)", "关系驱动、重信任、决策周期长",
         "WhatsApp 普及度高、面谈重要性大",
         "关系建立在前、价格谈判在后、可议价",
         "L/C 为主、部分 TT",
         "避免涉及宗教话题、左手递物不妥",
         "送礼常见、咖啡/椰枣是文化符号",
         "工作时间 9:00-14:00 + 16:00-19:00（斋月不同）"],
        ["东南亚", "华人网络、灵活、价格敏感",
         "微信/WhatsApp 均可、线上沟通高效",
         "关系导向、可议价空间大",
         "TT 为主",
         "尊重当地宗教信仰（伊斯兰/佛教/天主教）",
         "小礼物有加分、注重面子文化",
         "工作时间 8:00-17:00 当地"],
    ]
    for ri, row in enumerate(cultures, 4):
        for ci, val in enumerate(row, 1):
            ws_.cell(row=ri, column=ci, value=val)
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for ri in range(7, 15):
        _style_row(ws_, ri, len(headers), alt=ri % 2 == 0)

    for c in range(1, len(headers) + 1):
        ws_.column_dimensions[get_column_letter(c)].width = 28


# ═══════════════════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════════════════

def generate_templates() -> Workbook:
    """生成包含 10 张管理表格的完整 Excel 工作簿。"""
    wb = Workbook()
    _build_selection(wb)
    _build_persona(wb)
    _build_market(wb)
    _build_channel(wb)
    _build_content(wb)
    _build_followup(wb)
    _build_pipeline(wb)
    _build_kpi(wb)
    _build_competitor(wb)
    _build_culture(wb)
    return wb
